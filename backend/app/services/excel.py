"""Excel import/export helpers."""

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.core.logging import logger
from app.db.models.order import OrderItem
from app.db.models.packing_record import PackingRecord
from app.db.models.product import Product
from app.db.models.service import Service

EXPORT_COLUMNS = [
    "Название",
    "Бренд",
    "Размер",
    "Цвет",
    "Баркод",
    "Артикул WB",
    "Ссылка WB",
    "ТЗ упаковка",
    "Поставщик",
    "Остаток",
]
REQUIRED_COLUMNS = {"Название", "Баркод", "Артикул WB", "Поставщик"}

RECEIVING_COLUMNS = [
    "Баркод",
    "Название товара",
    "Дата приемки",
    "Кол-во план",
    "Кол-во факт",
    "Расхождения",
    "Комментарии",
]

SERVICES_COLUMNS = ["Категория", "Название", "Цена", "Ед.", "Комментарий", "Активна"]

FBO_COLUMNS = [
    "ID сотрудника",
    "Номер палета",
    "Номер короба",
    "Баркод",
    "Название товара",
    "Кол-во",
    "Склад",
    "Баркод короба",
    "Дата поставки",
    "packing_id",
]

ORDER_IMPORT_COLUMNS = [
    "Название",
    "Бренд",
    "Размер",
    "Цвет",
    "Баркод",
    "Артикул WB",
    "Ссылка WB",
    "ТЗ упаковка",
    "Поставщик",
    "Количество",
]
ORDER_IMPORT_REQUIRED = {"Название", "Количество"}

# Column widths (character units) for each export, in same order as column list.
ORDER_IMPORT_WIDTHS = [35, 18, 12, 12, 18, 16, 45, 25, 20, 12]
EXPORT_WIDTHS = [35, 18, 12, 12, 18, 16, 45, 25, 20, 10]
RECEIVING_WIDTHS = [18, 35, 14, 12, 12, 12, 30]
FBO_WIDTHS = [14, 14, 14, 18, 35, 8, 20, 22, 14, 12]
FBO_SUPPLY_BOX_WIDTHS = [14, 22]
SERVICES_WIDTHS = [18, 30, 10, 8, 35, 8]


def _set_excel_column_widths(buffer: BytesIO, widths: list[int]) -> BytesIO:
    """Set column widths on the first sheet. Returns new BytesIO with updated workbook."""
    buffer.seek(0)
    wb = load_workbook(buffer, read_only=False)
    ws = wb.active
    if ws is not None:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = min(max(w, 8), 55)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_order_items(order_items: list[OrderItem]) -> BytesIO:
    """Export order items to Excel (for import template / export)."""
    try:
        rows = []
        for item in order_items:
            product = item.product
            rows.append(
                {
                    "Название": product.name if product else "",
                    "Бренд": product.brand if product else "",
                    "Размер": product.size if product else "",
                    "Цвет": product.color if product else "",
                    "Баркод": product.barcode if product else "",
                    "Артикул WB": product.wb_article if product else "",
                    "Ссылка WB": product.wb_url if product else "",
                    "ТЗ упаковка": product.packing_instructions if product else "",
                    "Поставщик": product.supplier_name if product else "",
                    "Количество": item.planned_qty or 0,
                }
            )
        df = pd.DataFrame(rows, columns=ORDER_IMPORT_COLUMNS)
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return _set_excel_column_widths(buffer, ORDER_IMPORT_WIDTHS)
    except Exception as exc:
        logger.exception("excel_export_order_items_failed", error=str(exc))
        raise


def export_order_items_template() -> BytesIO:
    """Export empty order items template."""
    buffer = BytesIO()
    pd.DataFrame(columns=ORDER_IMPORT_COLUMNS).to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return _set_excel_column_widths(buffer, ORDER_IMPORT_WIDTHS)


def parse_orders_excel(file_bytes: bytes) -> list[dict]:
    """Parse order items from Excel (Название, Количество required). Returns list of dicts with product fields + planned_qty."""
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        missing = ORDER_IMPORT_REQUIRED.difference(set(df.columns))
        if missing:
            raise ValueError(f"Отсутствуют столбцы: {', '.join(sorted(missing))}")
        df = df.fillna("")
        result = []
        for _, row in df.iterrows():
            name = str(row.get("Название", "")).strip()
            qty_val = row.get("Количество", 0)
            try:
                planned_qty = int(qty_val) if qty_val != "" else 0
            except (TypeError, ValueError):
                planned_qty = 0
            if not name or planned_qty <= 0:
                continue
            result.append(
                {
                    "name": name,
                    "brand": str(row.get("Бренд", "")).strip() or None,
                    "size": str(row.get("Размер", "")).strip() or None,
                    "color": str(row.get("Цвет", "")).strip() or None,
                    "barcode": str(row.get("Баркод", "")).strip() or None,
                    "wb_article": str(row.get("Артикул WB", "")).strip() or None,
                    "wb_url": str(row.get("Ссылка WB", "")).strip() or None,
                    "packing_instructions": str(row.get("ТЗ упаковка", "")).strip() or None,
                    "supplier_name": str(row.get("Поставщик", "")).strip() or None,
                    "planned_qty": planned_qty,
                }
            )
        return result
    except Exception as exc:
        logger.exception("excel_parse_orders_failed", error=str(exc))
        raise


def export_products_template() -> BytesIO:
    """Export empty Excel template with required columns."""
    buffer = BytesIO()
    pd.DataFrame(columns=EXPORT_COLUMNS).to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)
    return _set_excel_column_widths(buffer, EXPORT_WIDTHS)


def export_products(products: list[Product]) -> BytesIO:
    """Export products to Excel in-memory file."""
    try:
        rows = []
        for product in products:
            rows.append(
                {
                    "Название": product.name,
                    "Бренд": product.brand,
                    "Размер": product.size,
                    "Цвет": product.color,
                    "Баркод": product.barcode,
                    "Артикул WB": product.wb_article,
                    "Ссылка WB": product.wb_url,
                    "ТЗ упаковка": product.packing_instructions,
                    "Поставщик": product.supplier_name,
                    "Остаток": product.stock_quantity,
                }
            )
        df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return _set_excel_column_widths(buffer, EXPORT_WIDTHS)
    except Exception as exc:
        logger.exception("excel_export_failed", error=str(exc))
        raise


def export_receiving(order_items: list[OrderItem]) -> BytesIO:
    """Export receiving data (order items) to Excel."""
    try:
        rows = []
        for item in order_items:
            product = item.product
            order = item.order
            diff = (item.planned_qty or 0) - (item.received_qty or 0)
            rows.append(
                {
                    "Баркод": product.barcode if product else "",
                    "Название товара": product.name if product else "",
                    "Дата приемки": order.updated_at.strftime("%d.%m.%Y") if order else "",
                    "Кол-во план": item.planned_qty or 0,
                    "Кол-во факт": item.received_qty or 0,
                    "Расхождения": diff,
                    "Комментарии": item.adjustment_note or "",
                }
            )
        df = pd.DataFrame(rows, columns=RECEIVING_COLUMNS)
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return _set_excel_column_widths(buffer, RECEIVING_WIDTHS)
    except Exception as exc:
        logger.exception("excel_export_receiving_failed", error=str(exc))
        raise


def export_fbo_shipping(
    packing_records: list[PackingRecord],
    delivery_date: str | None = None,
) -> BytesIO:
    """Export FBO shipping (packing records) to Excel. delivery_date optional (e.g. from shipment)."""
    try:
        sorted_records = sorted(
            packing_records,
            key=lambda r: (
                r.pallet_number if r.pallet_number is not None else 0,
                r.box_number if r.box_number is not None else 0,
            ),
        )
        rows = []
        for rec in sorted_records:
            product = rec.product
            employee = rec.employee
            rows.append(
                {
                    "ID сотрудника": employee.employee_code if employee else "",
                    "Номер палета": rec.pallet_number or "",
                    "Номер короба": rec.box_number or "",
                    "Баркод": product.barcode if product else "",
                    "Название товара": product.name if product else "",
                    "Кол-во": rec.quantity or 0,
                    "Склад": rec.warehouse or "",
                    "Баркод короба": rec.box_barcode or "",
                    "Дата поставки": delivery_date or "",
                    "packing_id": rec.id,
                }
            )
        df = pd.DataFrame(rows, columns=FBO_COLUMNS)
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return _set_excel_column_widths(buffer, FBO_WIDTHS)
    except Exception as exc:
        logger.exception("excel_export_fbo_failed", error=str(exc))
        raise


def parse_fbo_excel(file_bytes: bytes) -> list[dict]:
    """Parse FBO Excel (packing_id, Баркод короба, Склад, Дата поставки)."""
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        df = df.fillna("")
        if "packing_id" not in df.columns:
            raise ValueError("В файле должен быть столбец packing_id")
        result = []
        for _, row in df.iterrows():
            pid_val = row.get("packing_id", "")
            try:
                packing_id = int(pid_val) if pid_val != "" else None
            except (TypeError, ValueError):
                packing_id = None
            if packing_id is None:
                continue

            box_barcode = str(row.get("Баркод короба", "")).strip() or None
            warehouse = str(row.get("Склад", "")).strip() or None

            delivery_date = row.get("Дата поставки", "")
            if isinstance(delivery_date, pd.Timestamp):
                delivery_date = delivery_date.strftime("%Y-%m-%d")
            else:
                delivery_date = str(delivery_date).strip() or None

            result.append(
                {
                    "packing_id": packing_id,
                    "box_barcode": box_barcode,
                    "warehouse": warehouse,
                    "delivery_date": delivery_date,
                }
            )
        return result
    except Exception as exc:
        logger.exception("excel_parse_fbo_failed", error=str(exc))
        raise


FBO_SUPPLY_BOX_COLUMNS = ["Номер короба", "Штрихкод"]


def export_fbo_supply_boxes(boxes: list) -> BytesIO:
    """Export FBO supply boxes to Excel for manual edit/import. Each box: box_number, external_barcode."""
    try:
        rows = []
        for b in sorted(boxes, key=lambda x: getattr(x, "box_number", x.get("box_number", 0))):
            box_number = getattr(b, "box_number", b.get("box_number", 0))
            barcode = getattr(b, "external_barcode", None) or b.get("external_barcode") or b.get("Штрихкод") or ""
            rows.append({"Номер короба": box_number, "Штрихкод": barcode or ""})
        df = pd.DataFrame(rows, columns=FBO_SUPPLY_BOX_COLUMNS)
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return _set_excel_column_widths(buffer, FBO_SUPPLY_BOX_WIDTHS)
    except Exception as exc:
        logger.exception("excel_export_fbo_supply_boxes_failed", error=str(exc))
        raise


def parse_fbo_supply_excel(file_bytes: bytes) -> list[dict]:
    """Parse FBO supply Excel (Номер короба, Штрихкод). Returns list of {box_number, barcode}."""
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        df = df.fillna("")
        if "Номер короба" not in df.columns or "Штрихкод" not in df.columns:
            raise ValueError("В файле должны быть столбцы «Номер короба» и «Штрихкод»")
        result = []
        for _, row in df.iterrows():
            try:
                num = int(row.get("Номер короба", 0))
            except (TypeError, ValueError):
                continue
            barcode = str(row.get("Штрихкод", "")).strip() or None
            result.append({"box_number": num, "barcode": barcode})
        return result
    except Exception as exc:
        logger.exception("excel_parse_fbo_supply_failed", error=str(exc))
        raise


def parse_products_excel(file_bytes: bytes) -> list[dict]:
    """Parse products from Excel bytes."""
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        missing = REQUIRED_COLUMNS.difference(set(df.columns))
        if missing:
            raise ValueError(f"Missing columns: {', '.join(sorted(missing))}")
        df = df.fillna("")
        products = []
        for _, row in df.iterrows():
            products.append(
                {
                    "name": str(row.get("Название", "")).strip(),
                    "brand": str(row.get("Бренд", "")).strip() or None,
                    "size": str(row.get("Размер", "")).strip() or None,
                    "color": str(row.get("Цвет", "")).strip() or None,
                    "barcode": str(row.get("Баркод", "")).strip() or None,
                    "wb_article": str(row.get("Артикул WB", "")).strip() or None,
                    "wb_url": str(row.get("Ссылка WB", "")).strip() or None,
                    "packing_instructions": str(row.get("ТЗ упаковка", "")).strip() or None,
                    "supplier_name": str(row.get("Поставщик", "")).strip() or None,
                }
            )
        return products
    except Exception as exc:
        logger.exception("excel_parse_failed", error=str(exc))
        raise


def export_services(services: list[Service]) -> BytesIO:
    """Export services (pricing) to Excel in-memory file."""
    try:
        rows = []
        for s in services:
            rows.append(
                {
                    "Категория": s.category,
                    "Название": s.name,
                    "Цена": float(s.price),
                    "Ед.": s.unit,
                    "Комментарий": s.comment or "",
                    "Активна": "Да" if s.is_active else "Нет",
                }
            )
        df = pd.DataFrame(rows, columns=SERVICES_COLUMNS)
        buffer = BytesIO()
        df.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        return _set_excel_column_widths(buffer, SERVICES_WIDTHS)
    except Exception as exc:
        logger.exception("excel_export_services_failed", error=str(exc))
        raise


def parse_services_excel(file_bytes: bytes) -> list[dict]:
    """Parse services from Excel bytes. Columns: Категория, Название, Цена, Ед., Комментарий."""
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        required = {"Категория", "Название", "Цена"}
        missing = required.difference(set(df.columns))
        if missing:
            raise ValueError(f"Missing columns: {', '.join(sorted(missing))}")
        df = df.fillna("")
        services = []
        for _, row in df.iterrows():
            cat = str(row.get("Категория", "")).strip()
            name = str(row.get("Название", "")).strip()
            if not cat or not name:
                continue
            price_val = row.get("Цена", 0)
            try:
                price = float(price_val) if price_val != "" else 0.0
            except (TypeError, ValueError):
                price = 0.0
            unit = str(row.get("Ед.", "шт")).strip() or "шт"
            comment = str(row.get("Комментарий", "")).strip() or None
            services.append(
                {
                    "category": cat,
                    "name": name,
                    "price": price,
                    "unit": unit,
                    "comment": comment,
                }
            )
        return services
    except Exception as exc:
        logger.exception("excel_parse_services_failed", error=str(exc))
        raise
